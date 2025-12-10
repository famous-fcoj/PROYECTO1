import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from datetime import datetime, date
from decimal import Decimal
from django.db.models import Count, Q
from .models import OrdenTrabajo, Tarea, Repuesto, Insumo 
from django.template.loader import get_template
from xhtml2pdf import pisa

# --- HELPER DE DATOS ---
def safe_cast(data, field_name, cast_type):
    value = data.get(field_name)
    if value in (None, '', 'null'): return None
    if cast_type == date:
        try: return datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError: return None 
    if cast_type == datetime:
        try:
            dt = datetime.strptime(value, '%Y-%m-%d')
            return datetime(dt.year, dt.month, dt.day, 0, 0, 0)
        except ValueError: return None
    if cast_type == int:
        try: return int(value)
        except: return 0 
    if cast_type == Decimal:
        try: return Decimal(str(value)) 
        except: return Decimal(0)
    return str(value).strip()

# --- HELPER CENTRALIZADO PARA DIBUJAR EXCEL ---
# Esta función recibe una hoja de cálculo (ws) y un objeto OT, y dibuja el reporte.
# Se usa tanto para el reporte individual como para el masivo.
def dibujar_ot_en_excel(ws, ot):
    # Estilos
    COLOR_PRIMARY = "1F3B4B"; COLOR_HEADER_BG = "E2E8F0"; COLOR_TEXT_WHITE = "FFFFFF"
    font_title = Font(name='Arial', size=16, bold=True, color=COLOR_TEXT_WHITE)
    font_section = Font(name='Arial', size=11, bold=True, color=COLOR_PRIMARY)
    font_header_table = Font(name='Arial', size=10, bold=True, color=COLOR_TEXT_WHITE)
    font_label = Font(name='Arial', size=10, bold=True, color="333333")
    font_data = Font(name='Arial', size=10, color="000000")
    
    border_thick = Side(style='medium', color=COLOR_PRIMARY)
    border_thin = Side(style='thin', color="94A3B8")
    border_box = Border(left=border_thick, right=border_thick, top=border_thick, bottom=border_thick)
    border_cell = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    fill_title = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
    fill_header_table = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    fill_label_bg = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.column_dimensions['A'].width = 6; ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 35; ws.column_dimensions['D'].width = 25; ws.column_dimensions['E'].width = 35

    # Título
    ws.merge_cells('A1:E2')
    title_cell = ws['A1']
    title_cell.value = f"ORDEN DE TRABAJO N° {ot.ot}"
    title_cell.font = font_title; title_cell.fill = fill_title; title_cell.alignment = center; title_cell.border = border_box

    current_row = 4
    
    def draw_row(r, l1, v1, l2, v2):
        for col, val, style in [(2, l1, 'label'), (3, v1, 'data'), (4, l2, 'label'), (5, v2, 'data')]:
            c = ws.cell(row=r, column=col, value=val)
            c.border = border_cell
            if style == 'label': c.font = font_label; c.fill = fill_label_bg
            else: c.font = font_data; c.alignment = left

    f_creacion = ot.fecha_inicio.strftime('%d/%m/%Y') if ot.fecha_inicio else '-'
    f_plan = ot.fecha_planificada.strftime('%d/%m/%Y') if ot.fecha_planificada else '-'
    
    draw_row(current_row, "Fecha Emisión", f_creacion, "Equipo / Máquina", ot.maquina); current_row += 1
    
    ws.cell(row=current_row, column=2, value="Descripción").font = font_label
    ws.cell(row=current_row, column=2).fill = fill_label_bg
    ws.cell(row=current_row, column=2).border = border_cell
    ws.merge_cells(f'C{current_row}:E{current_row}')
    desc = ws.cell(row=current_row, column=3, value=ot.descripcion)
    desc.font = font_data; desc.alignment = left; desc.border = border_cell
    ws.cell(row=current_row, column=4).border = border_cell; ws.cell(row=current_row, column=5).border = border_cell
    current_row += 1

    draw_row(current_row, "Marca", ot.marca, "Modelo", ot.modelo); current_row += 1
    draw_row(current_row, "Ubicación", ot.ubicacion, "Tipo Acción", ot.tipo_accion); current_row += 1
    draw_row(current_row, "Odómetro", ot.odometro, "Responsable", ot.encargado); current_row += 1
    draw_row(current_row, "Supervisor", ot.supervisor, "Fecha Planificada", f_plan); current_row += 1
    draw_row(current_row, "Estado", ot.get_estado_display(), "Folio Interno", ot.ot); current_row += 2

    # Tablas
    def header_sec(row, title):
        ws.merge_cells(f'A{row}:E{row}')
        c = ws.cell(row=row, column=1, value=title)
        c.font = font_section; c.border = Border(bottom=border_thick)

    def table_head(row, headers):
        for idx, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=idx, value=h)
            c.font = font_header_table; c.fill = fill_header_table; c.alignment = center; c.border = border_cell

    # Tareas
    header_sec(current_row, "1. TAREAS A EJECUTAR"); current_row += 1
    table_head(current_row, ["#", "Detalle", "T. Est", "T. Real", "Check"]); current_row += 1
    
    tareas = ot.tareas.all()
    if not tareas:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws.cell(row=current_row, column=1, value="Sin tareas").alignment = center
        for i in range(1,6): ws.cell(row=current_row, column=i).border = border_cell
        current_row += 1
    else:
        for i, t in enumerate(tareas, 1):
            vals = [i, t.detalle, t.tiempo_estimado, t.tiempo_real, "[ ]"]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=current_row, column=col, value=v)
                c.alignment = left if col == 2 else center
                c.border = border_cell
            current_row += 1
    current_row += 1

    # Materiales
    header_sec(current_row, "2. MATERIALES"); current_row += 1
    table_head(current_row, ["#", "Código", "Descripción", "Cant", "Estado"]); current_row += 1
    materiales = list(ot.repuestos.all()) + list(ot.insumos.all())
    
    if not materiales:
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws.cell(row=current_row, column=1, value="Sin materiales").alignment = center
        for i in range(1,6): ws.cell(row=current_row, column=i).border = border_cell
        current_row += 1
    else:
        for i, m in enumerate(materiales, 1):
            vals = [i, m.codigo, m.descripcion, m.cantidad, ""]
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=current_row, column=col, value=v)
                c.alignment = left if col == 3 else center
                c.border = border_cell
            current_row += 1
    current_row += 2

    # Cierre
    ws.merge_cells(f'A{current_row}:E{current_row}')
    ws.cell(row=current_row, column=1, value="CIERRE").font = font_section; ws.cell(row=current_row, column=1).border = Border(bottom=border_thick)
    current_row += 1
    ws.merge_cells(f'A{current_row}:E{current_row+3}')
    obs = ws.cell(row=current_row, column=1, value=ot.observacion or "(Sin observaciones)")
    obs.alignment = Alignment(horizontal='left', vertical='top'); obs.border = border_box
    current_row += 5
    
    ws.cell(row=current_row, column=2, value="_________________").alignment = center
    ws.cell(row=current_row, column=4, value="_________________").alignment = center
    current_row += 1
    ws.cell(row=current_row, column=2, value=f"Rev: {ot.revisado_por or ''}").alignment = center
    ws.cell(row=current_row, column=4, value=f"Rec: {ot.recibido_por or ''}").alignment = center


# --- VISTAS API (Sin cambios mayores) ---
@csrf_exempt 
def recibir_orden_trabajo(request): 
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            tareas = data.pop('tareas', []); repuestos = data.pop('repuestos', []); insumos = data.pop('insumos', [])
            ot_data = {
                'ot': data.get('numero_ot'), 'encargado': data.get('responsable_ejecucion'),
                'maquina': data.get('equipo_maquina'), 'descripcion': data.get('descripcion'),
                'marca': data.get('marca'), 'ubicacion': data.get('ubicacion'),
                'modelo': data.get('modelo'), 'tipo_accion': data.get('tipo_accion'),
                'tipo_falla': data.get('tipo_accion'), 'odometro': safe_cast(data, 'odometro', int),
                'supervisor': data.get('supervisor'), 'fecha_planificada': safe_cast(data, 'fechaPlanificada', date),
                'fecha_inicio': safe_cast(data, 'fechaCreacion', datetime), 'revisado_por': data.get('revisadoPor'),
                'fecha_revision': safe_cast(data, 'fechaRevision', date), 'recibido_por': data.get('recibidoPor'),
                'observacion': data.get('observaciones'), 'estado': data.get('estado', 'PENDIENTE')
            }
            ot_data = {k: v for k, v in ot_data.items() if v is not None}
            ot_num = ot_data.pop('ot')
            ot_obj, created = OrdenTrabajo.objects.get_or_create(ot=ot_num, defaults=ot_data)
            if not created:
                for k, v in ot_data.items(): setattr(ot_obj, k, v)
                ot_obj.save()
            ot_obj.tareas.all().delete(); ot_obj.repuestos.all().delete(); ot_obj.insumos.all().delete()
            for i, t in enumerate(tareas, 1): Tarea.objects.create(orden=ot_obj, numero_item=i, **t)
            for i, r in enumerate(repuestos, 1): Repuesto.objects.create(orden=ot_obj, numero_item=i, **r)
            for i, ins in enumerate(insumos, 1): Insumo.objects.create(orden=ot_obj, numero_item=i, **ins)
            return JsonResponse({'status': 'success', 'message': 'Guardado OK', 'numero_ot': ot_obj.ot})
        except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'error': 'Method not allowed'}, status=405)

def lista_ots_api(request):
    ordenes = OrdenTrabajo.objects.all()
    data = [{'ot': o.ot, 'fecha': o.fecha_inicio.strftime('%Y-%m-%d') if o.fecha_inicio else '-', 'maquina': o.maquina, 'encargado': o.encargado, 'estado': o.estado} for o in ordenes]
    try: data.sort(key=lambda x: int(x['ot'].split('-')[1]))
    except: data.sort(key=lambda x: x['ot'])
    return JsonResponse({'ordenes': data})

def detalle_ot_api(request, ot_num):
    ot = get_object_or_404(OrdenTrabajo, ot=ot_num)
    data = {
        'numero_ot': ot.ot, 'fechaCreacion': ot.fecha_inicio.strftime('%Y-%m-%d') if ot.fecha_inicio else '',
        'equipo_maquina': ot.maquina, 'descripcion': ot.descripcion, 'marca': ot.marca,
        'ubicacion': ot.ubicacion, 'modelo': ot.modelo, 'tipo_accion': ot.tipo_accion,
        'odometro': ot.odometro, 'responsable_ejecucion': ot.encargado,
        'fechaPlanificada': ot.fecha_planificada.strftime('%Y-%m-%d') if ot.fecha_planificada else '',
        'supervisor': ot.supervisor, 'revisadoPor': ot.revisado_por,
        'fechaRevision': ot.fecha_revision.strftime('%Y-%m-%d') if ot.fecha_revision else '',
        'recibidoPor': ot.recibido_por, 'observaciones': ot.observacion, 'estado': ot.estado,
        'tareas': list(ot.tareas.values('detalle', 'tiempo_estimado', 'tiempo_real')),
        'repuestos': list(ot.repuestos.values('codigo', 'descripcion', 'cantidad')),
        'insumos': list(ot.insumos.values('codigo', 'descripcion', 'cantidad'))
    }
    return JsonResponse(data)

@csrf_exempt
def eliminar_ot_api(request, ot_num):
    if request.method == 'POST':
        try: OrdenTrabajo.objects.get(ot=ot_num).delete(); return JsonResponse({'status': 'success'})
        except: return JsonResponse({'status': 'error'}, status=404)
    return JsonResponse({'status': 'error'}, status=400)

def ot_resumen_api(request):
    total = OrdenTrabajo.objects.count(); pendientes = OrdenTrabajo.objects.filter(estado='PENDIENTE').count()
    proceso = OrdenTrabajo.objects.filter(estado='EN_PROCESO').count(); finalizadas = OrdenTrabajo.objects.filter(estado='FINALIZADA').count()
    all_types = OrdenTrabajo.objects.values('tipo_falla').annotate(total=Count('id')).order_by('-total')
    top_types = list(all_types[:5]); rest = sum(i['total'] for i in all_types[5:])
    if rest > 0: top_types.append({'tipo_falla': 'Otros', 'total': rest})
    carga = OrdenTrabajo.objects.values('encargado').annotate(total=Count('id'), pendientes=Count('id', filter=Q(estado='PENDIENTE')), proceso=Count('id', filter=Q(estado='EN_PROCESO')), finalizadas=Count('id', filter=Q(estado='FINALIZADA'))).order_by('-total')[:7]
    maquina = OrdenTrabajo.objects.values('maquina').annotate(total=Count('id')).order_by('-total')[:5]
    estado = OrdenTrabajo.objects.values('estado').annotate(total=Count('id'))
    return JsonResponse({'resumen_estado': list(estado), 'resumen_maquina': list(maquina), 'resumen_tipo': top_types, 'carga_laboral': list(carga), 'kpis': {'total': total, 'pendientes': pendientes, 'proceso': proceso, 'finalizadas': finalizadas}})

def siguiente_folio_api(request):
    num = 1
    while True:
        if not OrdenTrabajo.objects.filter(ot=f'OT-{num}').exists(): break
        num += 1
    return JsonResponse({'nuevo_folio': f'OT-{num}'})

# --- EXPORTACIONES INDIVIDUALES ---
def exportar_ot_excel(request, ot_num):
    ot = get_object_or_404(OrdenTrabajo, ot=ot_num)
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = f"OT_{ot.ot}"
    # Usamos el helper centralizado
    dibujar_ot_en_excel(ws, ot)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{ot.ot.replace("-","_")}.xlsx"'
    wb.save(response); return response

def exportar_ot_pdf(request, ot_num):
    ot = get_object_or_404(OrdenTrabajo, ot=ot_num)
    template = get_template('ot_pdf.html')
    html = template.render({'ot': ot})
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{ot.ot.replace("-","_")}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err: return HttpResponse('Error PDF', status=500)
    return response

# --- NUEVAS EXPORTACIONES MASIVAS ---
def exportar_todas_excel(request):
    wb = openpyxl.Workbook()
    # Borrar la hoja por defecto que crea openpyxl
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Obtener todas las OTs y ordenarlas numéricamente
    ordenes = list(OrdenTrabajo.objects.all())
    try: ordenes.sort(key=lambda x: int(x.ot.split('-')[1]))
    except: ordenes.sort(key=lambda x: x.ot)

    if not ordenes:
        ws = wb.create_sheet(title="Sin Registros")
        ws['A1'] = "No hay órdenes de trabajo para exportar."
    else:
        # Iterar y crear una hoja por cada OT
        for ot in ordenes:
            # El nombre de la hoja no puede exceder 31 caracteres
            sheet_title = f"{ot.ot}"[:31]
            ws = wb.create_sheet(title=sheet_title)
            # Reutilizamos la misma función de dibujo
            dibujar_ot_en_excel(ws, ot)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Completo_OTs.xlsx"'
    wb.save(response)
    return response

def exportar_todas_pdf(request):
    # Obtener y ordenar
    ordenes = list(OrdenTrabajo.objects.all())
    try: ordenes.sort(key=lambda x: int(x.ot.split('-')[1]))
    except: ordenes.sort(key=lambda x: x.ot)

    if not ordenes:
        return HttpResponse("No hay órdenes de trabajo para exportar.", content_type='text/plain')

    # Usamos el NUEVO template masivo que itera internamente
    template = get_template('ot_masivo_pdf.html')
    # Le pasamos la lista completa de órdenes
    html = template.render({'ordenes': ordenes})
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Completo_OTs.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err: return HttpResponse('Error generando el PDF masivo', status=500)
    return response